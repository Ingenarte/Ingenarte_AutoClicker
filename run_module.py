import customtkinter as ctk
import time 
import pyautogui
import pyperclip
from pynput import mouse
import threading
import json
import tkinter as tk
from tkinter import filedialog
import datetime
import os
import subprocess
import sys
from openpyxl import load_workbook  
from modals import recursivity_modal
from modals.modal_input import open_input_modal 
from modals.image_modal import open_image_modal
from modals.data_modal import open_data_modal
from image_engine import find_image

import logging
logging.getLogger("PIL").setLevel(logging.WARNING)
logging.getLogger("PIL.PngImagePlugin").setLevel(logging.WARNING)

# -----------------------------
# Global configuration (will be loaded from run.json if it exists)
# -----------------------------
global_config = {}

# -----------------------------
# Logging function: prints and appends to run-log.txt
# -----------------------------
def log_action(message):
    """Append a log message with timestamp to run-log.txt and print it."""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d | %H:%M:%S")
    log_line = f"{timestamp} | {message}"
    with open("run-log.txt", "a", encoding="utf-8") as log_file:
        log_file.write(log_line + "\n")
    print(log_line,flush=True)

# -----------------------------
# Processing functions for steps
# -----------------------------

def get_clicks(value_str):
    """Convert a numeric string or word to an integer."""
    try:
        # Try to convert directly if it's a numeric string
        return int(value_str)
    except ValueError:
        # Otherwise, map known number words to integers
        mapping = {
            "one": 1,
            "two": 2,
            "three": 3,
            "four": 4,
            "five": 5,
            "six": 6,
            "seven": 7,
            "eight": 8,
            "nine": 9,
            "ten": 10
        }
        return mapping.get(value_str.lower(), 1)














def process_input_step(step_config, step_number, tab_name):
    """
    Processes an input step for either mouse or keyboard actions.
    
    For mouse input:
      • Uses `position` ("100x200") or falls back to current pointer.
      • Optionally moves by `mouse_movement` ("dx×dy") before clicking.
      • Clicks `mouse_click_qty` times with the specified button.

    For keyboard input:
      • If 'keyboard_ascii' contains " + ", holds modifiers and presses the rest.
      • Repeats the press sequence `keyboard_repeat` times (default 1).

    Finally sleeps for `input_sleep` seconds if given.
    """


    ### Macos Double Triple Clicker 

    import sys
    import time
    import pyautogui

    if sys.platform == "darwin":
        import Quartz

    def double_click(x, y, interval=0.2, button='left'):
        btn_q = Quartz.kCGMouseButtonLeft if button == 'left' else Quartz.kCGMouseButtonRight
        if sys.platform == "darwin":
            # Primer clic (clickCount = 1)
            down1 = Quartz.CGEventCreateMouseEvent(
                None,
                Quartz.kCGEventLeftMouseDown if btn_q == Quartz.kCGMouseButtonLeft else Quartz.kCGEventRightMouseDown,
                (x, y),
                btn_q
            )
            Quartz.CGEventSetIntegerValueField(down1, Quartz.kCGMouseEventClickState, 1)
            Quartz.CGEventPost(Quartz.kCGHIDEventTap, down1)

            up1 = Quartz.CGEventCreateMouseEvent(
                None,
                Quartz.kCGEventLeftMouseUp if btn_q == Quartz.kCGMouseButtonLeft else Quartz.kCGEventRightMouseUp,
                (x, y),
                btn_q
            )
            Quartz.CGEventSetIntegerValueField(up1, Quartz.kCGMouseEventClickState, 1)
            Quartz.CGEventPost(Quartz.kCGHIDEventTap, up1)

            time.sleep(interval)

            # Segundo clic (clickCount = 2)
            down2 = Quartz.CGEventCreateMouseEvent(
                None,
                Quartz.kCGEventLeftMouseDown if btn_q == Quartz.kCGMouseButtonLeft else Quartz.kCGEventRightMouseDown,
                (x, y),
                btn_q
            )
            Quartz.CGEventSetIntegerValueField(down2, Quartz.kCGMouseEventClickState, 2)
            Quartz.CGEventPost(Quartz.kCGHIDEventTap, down2)

            up2 = Quartz.CGEventCreateMouseEvent(
                None,
                Quartz.kCGEventLeftMouseUp if btn_q == Quartz.kCGMouseButtonLeft else Quartz.kCGEventRightMouseUp,
                (x, y),
                btn_q
            )
            Quartz.CGEventSetIntegerValueField(up2, Quartz.kCGMouseEventClickState, 2)
            Quartz.CGEventPost(Quartz.kCGHIDEventTap, up2)
        else:
            pyautogui.click(x=x, y=y, clicks=2, interval=interval, button=button)

    def triple_click(x, y, interval=0.2, button='left'):
        btn_q = Quartz.kCGMouseButtonLeft if button == 'left' else Quartz.kCGMouseButtonRight
        if sys.platform == "darwin":
            # Clic 1 (clickCount = 1)
            for state in (1, 2, 3):
                down = Quartz.CGEventCreateMouseEvent(
                    None,
                    Quartz.kCGEventLeftMouseDown if btn_q == Quartz.kCGMouseButtonLeft else Quartz.kCGEventRightMouseDown,
                    (x, y),
                    btn_q
                )
                Quartz.CGEventSetIntegerValueField(down, Quartz.kCGMouseEventClickState, state)
                Quartz.CGEventPost(Quartz.kCGHIDEventTap, down)

                up = Quartz.CGEventCreateMouseEvent(
                    None,
                    Quartz.kCGEventLeftMouseUp if btn_q == Quartz.kCGMouseButtonLeft else Quartz.kCGEventRightMouseUp,
                    (x, y),
                    btn_q
                )
                Quartz.CGEventSetIntegerValueField(up, Quartz.kCGMouseEventClickState, state)
                Quartz.CGEventPost(Quartz.kCGHIDEventTap, up)

                if state < 3:
                    time.sleep(interval)
        else:
            pyautogui.click(x=x, y=y, clicks=3, interval=interval, button=button)

    ### --------- Macos Double Triple Clicker 








    action = step_config.get("input", {})
    log_action(f"[Tab {tab_name} | Step {step_number}] Processing input: {action}")
    src = action.get("input_from", "").strip().lower()

    # ─── MOUSE ───────────────────────────────────────────────────────────────
    if src == "mouse":
        pos = step_config.get("position", "").strip()
        if pos:
            try:
                x_str, y_str = pos.split("x")
                x, y = int(x_str), int(y_str)
            except:
                x, y = pyautogui.position()
                log_action(f"⚠️ Bad position '{pos}', using current mouse ({x},{y})")
        else:
            x, y = pyautogui.position()
            log_action(f"⚠️ No position specified – using current mouse ({x},{y})")

        # optional movement offset "dx×dy"
        mv = action.get("mouse_movement", "").strip()
        if mv:
            try:
                if "×" in mv:
                    dx, dy = map(int, mv.split("×"))
                elif "x" in mv:
                    dx, dy = map(int, mv.split("x"))
                else:
                    # no separator → horizontal move only
                    dx, dy = int(mv), 0

                x += dx
                y += dy
                log_action(f"Moved by offset ({dx},{dy}) → ({x},{y})")
            except Exception:
                log_action(f"⚠️ Invalid mouse_movement '{mv}' – skipping offset")

        btn = action.get("mouse_event", "Left").strip().lower()
        if btn not in ("left", "right", "middle"):
            log_action(f"⚠️ Invalid mouse button '{btn}', defaulting to 'left'")
            btn = "left"

        clicks = get_clicks(action.get("mouse_click_qty", "1"))

        # 5) Leer intervalo entre clics (por defecto 0.2 s)
        try:
            intervalo = float(action.get("click_interval", "0.2") or 0.2)
        except:
            intervalo = 0.2

        # 6) Definición de las funciones “double_click” y “triple_click” localmente,
        #    pero sólo importamos/llamamos a Quartz si estamos en macOS:
        if sys.platform == "darwin":
            import Quartz

            def double_click(x0, y0, interval=0.2, button='left'):
                btn_q = Quartz.kCGMouseButtonLeft if button == 'left' else Quartz.kCGMouseButtonRight

                # 1er clic (clickCount = 1)
                e_down1 = Quartz.CGEventCreateMouseEvent(
                    None,
                    Quartz.kCGEventLeftMouseDown if btn_q == Quartz.kCGMouseButtonLeft else Quartz.kCGEventRightMouseDown,
                    (x0, y0),
                    btn_q
                )
                Quartz.CGEventSetIntegerValueField(e_down1, Quartz.kCGMouseEventClickState, 1)
                Quartz.CGEventPost(Quartz.kCGHIDEventTap, e_down1)

                e_up1 = Quartz.CGEventCreateMouseEvent(
                    None,
                    Quartz.kCGEventLeftMouseUp if btn_q == Quartz.kCGMouseButtonLeft else Quartz.kCGEventRightMouseUp,
                    (x0, y0),
                    btn_q
                )
                Quartz.CGEventSetIntegerValueField(e_up1, Quartz.kCGMouseEventClickState, 1)
                Quartz.CGEventPost(Quartz.kCGHIDEventTap, e_up1)

                time.sleep(interval)

                # 2º clic (clickCount = 2)
                e_down2 = Quartz.CGEventCreateMouseEvent(
                    None,
                    Quartz.kCGEventLeftMouseDown if btn_q == Quartz.kCGMouseButtonLeft else Quartz.kCGEventRightMouseDown,
                    (x0, y0),
                    btn_q
                )
                Quartz.CGEventSetIntegerValueField(e_down2, Quartz.kCGMouseEventClickState, 2)
                Quartz.CGEventPost(Quartz.kCGHIDEventTap, e_down2)

                e_up2 = Quartz.CGEventCreateMouseEvent(
                    None,
                    Quartz.kCGEventLeftMouseUp if btn_q == Quartz.kCGMouseButtonLeft else Quartz.kCGEventRightMouseUp,
                    (x0, y0),
                    btn_q
                )
                Quartz.CGEventSetIntegerValueField(e_up2, Quartz.kCGMouseEventClickState, 2)
                Quartz.CGEventPost(Quartz.kCGHIDEventTap, e_up2)

            def triple_click(x0, y0, interval=0.2, button='left'):
                btn_q = Quartz.kCGMouseButtonLeft if button == 'left' else Quartz.kCGMouseButtonRight
                for state in (1, 2, 3):
                    e_down = Quartz.CGEventCreateMouseEvent(
                        None,
                        Quartz.kCGEventLeftMouseDown if btn_q == Quartz.kCGMouseButtonLeft else Quartz.kCGEventRightMouseDown,
                        (x0, y0),
                        btn_q
                    )
                    Quartz.CGEventSetIntegerValueField(e_down, Quartz.kCGMouseEventClickState, state)
                    Quartz.CGEventPost(Quartz.kCGHIDEventTap, e_down)

                    e_up = Quartz.CGEventCreateMouseEvent(
                        None,
                        Quartz.kCGEventLeftMouseUp if btn_q == Quartz.kCGMouseButtonLeft else Quartz.kCGEventRightMouseUp,
                        (x0, y0),
                        btn_q
                    )
                    Quartz.CGEventSetIntegerValueField(e_up, Quartz.kCGMouseEventClickState, state)
                    Quartz.CGEventPost(Quartz.kCGHIDEventTap, e_up)

                    if state < 3:
                        time.sleep(interval)
        else:
            # En Windows/Linux no necesitamos Quartz, PyAutoGUI lo maneja
            def double_click(x0, y0, interval=0.2, button='left'):
                pyautogui.click(x=x0, y=y0, clicks=2, interval=interval, button=button)

            def triple_click(x0, y0, interval=0.2, button='left'):
                pyautogui.click(x=x0, y=y0, clicks=3, interval=interval, button=button)

        # 7) Llamada al método correcto según número de clics:
        if clicks == 2:
            double_click(x, y, interval=intervalo, button=btn)
            log_action(f"Mouse double‐clicked at ({x},{y}) interval={intervalo}s")
        elif clicks == 3:
            triple_click(x, y, interval=intervalo, button=btn)
            log_action(f"Mouse triple‐clicked at ({x},{y}) interval={intervalo}s")
        else:
            pyautogui.click(x=x, y=y, clicks=clicks, interval=intervalo, button=btn)
            log_action(f"Mouse clicked at ({x},{y}) {clicks}× {btn} interval={intervalo}s")


    # ─── KEYBOARD ────────────────────────────────────────────────────────────
    elif src == "keyboard":
        text = action.get("keyboard_ascii", "").strip()

        # clamp repeat to ≥1
        try:
            repeat = int(action.get("keyboard_repeat", "") or 1)
        except:
            repeat = 1
        if repeat < 1:
            repeat = 1

        if " + " in text:
            parts = [k.strip() for k in text.split(" + ")]
            modifiers, keys = [], []
            for k in parts:
                kl = k.lower()
                if kl in ("shift","ctrl","alt","command","meta"):
                    modifiers.append("command" if kl in ("command","meta") else kl)
                else:
                    keys.append(kl)

            log_action(f"Key combo {text} ×{repeat}")
            for _ in range(repeat):
                for mod in modifiers:
                    pyautogui.keyDown(mod)
                for k in keys:
                    pyautogui.press(k)
                    time.sleep(0.2)
                for mod in modifiers:
                    pyautogui.keyUp(mod)

        else:
            key = text.lower()
            log_action(f"Key press {key} ×{repeat}")
            for _ in range(repeat):
                pyautogui.press(key)

    else:
        log_action("Input step: Unknown input_from, skipping.")

    # ─── SLEEP ────────────────────────────────────────────────────────────────
    try:
        sleep_after = float(action.get("input_sleep", "0") or 0)
    except:
        sleep_after = 0
    if sleep_after > 0:
        log_action(f"Waiting {sleep_after}s after input.")
        time.sleep(sleep_after)


def process_image_step(step_config, step_number, tab_name):
    """
    Executes one “image” step using our new image_engine:
      • Captures a debug screenshot via PyAutoGUI (all-monitors stitch).
      • Calls find_image() to locate the template on any monitor.
      • Optionally waits before clicking, then moves & clicks.
      • Sleeps afterwards if requested.
    """
    # 1) debug screenshot
    # dbg = pyautogui.screenshot()
    # dbg.save("debug_screenshot.png")

    action = step_config.get("image", {})
    log_action(f"[Tab {tab_name} | Step {step_number}] IMAGE → {action}")

    img_path = action.get("image_path", "")
    if not img_path:
        log_action("⚠️ Image step: No image_path specified.")
        return

    # parse params
    try:
        conf = float(action.get("image_confidence", 1.0))
    except:
        log_action("⚠️ Invalid image_confidence; defaulting to 1.0")
        conf = 1.0

    try:
        to = float(action.get("image_timeout", 0))
    except:
        log_action("⚠️ Invalid image_timeout; defaulting to 0")
        to = 0

    # 2) locate via our engine
    loc = find_image(
        img_path,
        base_confidence=conf,
        timeout=None if to <= 0 else to,
        poll_every=0.5,
        max_attempts=20
    )
    if not loc:
        log_action(f"⚠️ Image not found within timeout ({to}s) at confidence≥{conf}")
        return

    x, y, w, h, mon_idx = loc
    log_action(f"✅ Found on monitor {mon_idx} → logical=({x},{y}) size=({w}×{h})")

    # 3) optional pre-click wait
    if action.get("image_wait", False):
        try:
            ws = float(action.get("image_wait_sleep", 0))
        except:
            ws = 0
        log_action(f"Waiting {ws}s before click…")
        time.sleep(ws)

    # 4) move & click
    center_x = x + w//2
    center_y = y + h//2
    pyautogui.moveTo(center_x, center_y)
    log_action(f"Moved mouse to ({center_x},{center_y})")

    if action.get("image_click", False):
        btn = action.get("image_click_LR", "Left").lower()
        btn = "left" if btn == "left" else "right"
        try:
            clicks = int(action.get("mouse_click_qty", 1))
        except:
            clicks = get_clicks(action.get("mouse_click_qty", "1"))
        pyautogui.click(clicks=clicks, button=btn)
        log_action(f"Clicked {clicks}× {btn} at ({center_x},{center_y})")
    else:
        log_action("❎ image_click flag is False — no click performed.")

    # 5) post-click sleep
    try:
        ss = float(action.get("image_sleep", 0))
    except:
        ss = 0
    if ss > 0:
        log_action(f"Sleeping {ss}s after click…")
        time.sleep(ss)

def process_data_step(step_config, step_number, tab_name):
    """
    Process a 'data' step by operating on an Excel file.
    Expected configuration keys:
      - data_path: The Excel file path.
      - data_cell: The cell (e.g., "A3") to operate on.
      - data_copy_paste: If equal to "Paste To" (case-insensitive), the clipboard contents are pasted into the cell.
                         If equal to "Copy To", the cell value is copied to the clipboard.
      - data_select_all: Boolean; if True, simulate a 'select all' keystroke.
      - Optionally, data_sheet to specify the sheet name. If not provided, the first sheet is used.
    """
    action = step_config.get("data", {})
    log_action(f"[Tab {tab_name} | Step {step_number}] Processing data: {action}")

    file_path = action.get("data_path", "")
    cell = action.get("data_cell", "")
    copy_paste_action = action.get("data_copy_paste", "")
    select_all = action.get("data_select_all", False)
    sheet_name = action.get("data_sheet", None)

    if not file_path or not os.path.exists(file_path):
        log_action("Data step: Invalid or missing file path.")
        return

    try:
        wb = load_workbook(file_path)
        if sheet_name is None:
            # Default to the first sheet if none specified.
            sheet_name = wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            log_action(f"Sheet {sheet_name} not found in {file_path}.")
            return

        sheet = wb[sheet_name]

        if copy_paste_action.lower() == "paste to":
            # Paste clipboard content into the cell.
            clipboard_value = pyperclip.paste()
            sheet[cell].value = clipboard_value
            wb.save(file_path)
            log_action(f"Pasted clipboard value '{clipboard_value}' into cell {cell} in {sheet_name} of {file_path}.")
        elif copy_paste_action.lower() == "copy from":
            # Copy the cell's value to the clipboard.
            cell_value = sheet[cell].value
            pyperclip.copy(str(cell_value))
            log_action(f"Copied value '{cell_value}' from cell {cell} in {sheet_name} of {file_path} to clipboard.")
        else:
            log_action(f"Data step: Unknown data_copy_paste action '{copy_paste_action}'.")

        # If data_select_all is True, simulate a 'select all' keystroke.
        if select_all:
            if sys.platform == "darwin":
                pyautogui.hotkey("command", "a")
            else:
                pyautogui.hotkey("ctrl", "a")
            log_action("Executed select all command.")

    except Exception as e:
        log_action(f"Error processing data step: {e}")

def process_position_step(step_config, step_number, tab_name):
    pos = step_config.get("position", "")
    log_action(f"[Tab {tab_name} | Step {step_number}] Processing position: {pos}")
    if pos:
        try:
            x_str, y_str = pos.split("x")
            x, y = int(x_str), int(y_str)
            pyautogui.moveTo(x, y)
            log_action(f"Moved mouse to ({x}, {y}).")
        except Exception as e:
            log_action(f"Error processing position: {e}")
    else:
        log_action("Position step: No position provided.")

def process_step(tab_name, step_number, step_config):
    if "position" in step_config:
        process_position_step(step_config, step_number, tab_name)
    if "input" in step_config:
        process_input_step(step_config, step_number, tab_name)
    if "image" in step_config:
        process_image_step(step_config, step_number, tab_name)
    if "data" in step_config:
        process_data_step(step_config, step_number, tab_name)

def process_recursivity(rec_config):
    r_steps_str = rec_config.get("r_steps", "")
    r_repeat = rec_config.get("r_repeat", 1)
    if not r_steps_str:
        log_action("Recursivity: No steps specified.")
        return
    steps_to_repeat = [s.strip() for s in r_steps_str.split(",") if s.strip()]
    log_action(f"Recursivity: Repeating steps {', '.join(steps_to_repeat)} {r_repeat} times")
    for rep in range(1, r_repeat + 1):
        log_action(f"Recursivity iteration {rep} started.")
        for tab_name, steps in global_config.get("tab_n", {}).items():
            for step_key, step_config in steps.items():
                step_num = step_key.split("_")[1]
                if step_num in steps_to_repeat:
                    process_step(tab_name, step_num, step_config)
        log_action(f"Recursivity iteration {rep} completed.")

# -----------------------------
# Main processing function (run the script)
# -----------------------------
def run_script():
    log_action("RUN: Starting execution.")
    for tab_name in sorted(global_config.get("tab_n", {}).keys(), key=lambda t: int(t.split()[1])):
        log_action(f"RUN: Processing {tab_name}")
        steps = global_config["tab_n"][tab_name]
        for step_key in sorted(steps.keys(), key=lambda x: int(x.split('_')[1])):
            step_num = step_key.split("_")[1]
            step_config = steps[step_key]
            process_step(tab_name, step_num, step_config)
    if "recursivity" in global_config:
        process_recursivity(global_config["recursivity"])
    log_action("RUN: Execution completed.")

# -----------------------------
# "RUN" functionality: execute run_module.py synchronously (passing run.json as argument) and then rename run.json
# -----------------------------
def run_and_rename():
    log_action("RUN: Saving configuration to run.json...")
    run_filename = "run.json"
    try:
        with open(run_filename, "w", encoding="utf-8") as f:
            json.dump(global_config, f, ensure_ascii=False, indent=4, sort_keys=True)
        log_action(f"RUN: Configuration saved to {run_filename}.")
    except Exception as e:
        log_action(f"RUN: Error saving configuration: {e}")
        return

    log_action("RUN: Executing run_module.py with run.json as argument. This may take up to an hour...")
    try:
        # Call run_module.py with run.json as an argument.
        subprocess.run([sys.executable, "run_module.py", run_filename], check=True)
        log_action("RUN: run_module.py executed successfully.")
    except subprocess.CalledProcessError as e:
        log_action(f"RUN: Error executing run_module.py: {e}")
        return

    try:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        new_filename = f"run_{timestamp}.json"
        os.rename(run_filename, new_filename)
        log_action(f"RUN: Configuration file renamed to {new_filename}.")
    except Exception as e:
        log_action(f"RUN: Failed to rename run.json: {e}")

    run_script()

# -----------------------------
# Main entry point (for the run module)
# -----------------------------
def main():
    # If no command-line argument is provided, default to "run.json"
    if len(sys.argv) < 2:
        config_file = "run.json"
    else:
        config_file = sys.argv[1]
        
    try:
        with open(config_file, "r", encoding="utf-8") as f:
            config = json.load(f)
            global global_config
            global_config = config
        log_action(f"Loaded configuration from {config_file}")
    except Exception as e:
        log_action(f"Error loading configuration from {config_file}: {e}")
        sys.exit(1)
    
    # Now call the processing function to execute the steps.
    run_script()

if __name__ == "__main__":
    main()